# daily_report_bot.py

import pyautogui
import time
import os
from datetime import datetime
from openpyxl import Workbook

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 1

# -----------------------------
# Step 1: Open Chrome
# -----------------------------
print("Opening Chrome...")

pyautogui.hotkey('win', 'r')
time.sleep(1)

pyautogui.write('chrome')
pyautogui.press('enter')

time.sleep(3)

# -----------------------------
# Step 2: Open Public Gold Website
# -----------------------------
print("Opening Public Gold website...")

pyautogui.hotkey('ctrl', 't')
pyautogui.write('https://publicgold.com.my/')
pyautogui.press('enter')

time.sleep(10)

# -----------------------------
# Step 3: Gold Prices
# -----------------------------
# Update these values from website

gold_999 = "599"
gold_916 = "570"

# -----------------------------
# Step 4: Create Excel Report
# -----------------------------
print("Creating Excel report...")

now = datetime.now()

wb = Workbook()
ws = wb.active
ws.title = "Daily Report"

# Headers
ws['A1'] = 'Date & Time'
ws['B1'] = '999 Price'
ws['C1'] = '916 Price'
ws['D1'] = 'Comment'

# Data Row
ws['A2'] = now.strftime("%Y-%m-%d %H:%M:%S")
ws['B2'] = gold_999
ws['C2'] = gold_916
ws['D2'] = 'Daily Gold Rate Check'

# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 30

# Save Excel file
filename = f"daily_report_{now.strftime('%Y-%m-%d')}.xlsx"

wb.save(filename)

print("Excel saved:", filename)

# -----------------------------
# Step 5: Open Excel File
# -----------------------------
print("Opening Excel report...")

os.startfile(filename)

time.sleep(8)

# Maximize Excel window
pyautogui.hotkey('win', 'up')

time.sleep(2)

# -----------------------------
# Step 6: Take Screenshot
# -----------------------------
print("Taking screenshot...")

screenshot = pyautogui.screenshot()
screenshot.save("daily_report.png")

print("Screenshot saved as daily_report.png")

print("Task Completed Successfully")