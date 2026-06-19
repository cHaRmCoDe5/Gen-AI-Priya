# playwright_final_demo.py

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from datetime import datetime

print("Starting Cricket Score Report Bot...")

with sync_playwright() as p:

    browser = p.chromium.launch(
        headless=False,
        slow_mo=1000
    )

    page = browser.new_page()

    # Open Cricbuzz
    page.goto(
        "https://www.cricbuzz.com/",
        wait_until="domcontentloaded",
        timeout=60000
    )

    page.wait_for_timeout(5000)

    # Screenshot Home Page
    page.screenshot(path="cricbuzz_home.png")

    print("Cricbuzz opened successfully")

    # Click Matches Tab
    try:
        page.locator("text=Matches").first.click()
        print("Matches tab clicked")
    except Exception as e:
        print("Could not click Matches tab:", e)

    page.wait_for_timeout(5000)

    # Screenshot Matches Page
    page.screenshot(path="matches_page.png")

    # Extract page title
    title = page.title()

    # Extract visible text from page
    page_text = page.locator("body").inner_text()

    # Take only first few lines for report
    match_summary = "\n".join(page_text.split("\n")[:10])

    print("\n===== MATCH REPORT =====")
    print(match_summary)

    # -------------------------
    # Create Excel Report
    # -------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Cricket Report"

    ws["A1"] = "Date & Time"
    ws["B1"] = "Website"
    ws["C1"] = "Match Report"

    ws["A2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["B2"] = "Cricbuzz"
    ws["C2"] = match_summary

    # Adjust column width
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 100

    filename = f"cricket_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    wb.save(filename)

    print(f"\nExcel Report Saved: {filename}")

    # Screenshot of final page
    page.screenshot(path="final_cricket_report.png")

    print("Screenshot Saved: final_cricket_report.png")

    browser.close()

print("Task Completed Successfully")