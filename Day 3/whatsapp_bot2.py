from openpyxl import load_workbook, Workbook
from datetime import datetime
from pathlib import Path
import json
import random
import time
import os
import sys

BASE_DIR = Path(__file__).resolve().parent
PROFILE_DIR = BASE_DIR / "whatsapp_profile"
SCREENSHOTS_DIR = BASE_DIR / "screenshots"
# ==========================================


def random_delay():
    time.sleep(random.randint(2, 5))


def sanitize_filename(value):
    safe_chars = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 ._-"
    cleaned = "".join(c if c in safe_chars else "_" for c in value)
    return cleaned.strip().replace(" ", "_")


def find_search_box(page, timeout=10000):
    selectors = [
        'div[contenteditable="true"][data-tab="3"]',
        'div[contenteditable="true"][data-tab="1"]',
        'div[role="textbox"][contenteditable="true"]',
        'div[title*="Search"]',
        '[data-testid="chat-list-search"]',
        '//div[contains(@data-testid, "chat-search")]',
        '//div[contains(@class, "_2_1wd")]',
    ]
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=timeout)
            return loc
        except:
            continue
    try:
        loc = page.get_by_role("textbox")
        loc.wait_for(state="visible", timeout=timeout)
        return loc
    except:
        return None


def find_message_box(page, timeout=10000):
    selectors = [
        'div[contenteditable="true"][data-tab="10"]',
        'div[role="textbox"][data-tab="10"]',
        'div[role="textbox"]',
        'div[contenteditable="true"][data-tab]'
    ]
    for sel in selectors:
        try:
            loc = page.locator(sel).last
            loc.wait_for(state="visible", timeout=timeout)
            return loc
        except:
            continue
    try:
        loc = page.locator('div[contenteditable="true"]').last
        loc.wait_for(state="visible", timeout=timeout)
        return loc
    except:
        return None


def read_contacts():
    folder = BASE_DIR
    candidates = [
        folder / "contacts.xlsx",
        folder / "contacts.xlsx.xlsx",
        folder / "contacts.xls",
    ]

    existing = [path for path in candidates if path.exists()]
    if len(existing) == 1:
        contacts_file = existing[0]
        if contacts_file.name == "contacts.xlsx.xlsx":
            corrected = folder / "contacts.xlsx"
            if corrected.exists():
                contacts_file = corrected
            else:
                contacts_file = contacts_file.rename(corrected)
                print("Automatically renamed contacts.xlsx.xlsx to contacts.xlsx")
        elif contacts_file.name != "contacts.xlsx":
            print(f"Using contacts file: {contacts_file.name}")
    elif len(existing) > 1:
        print("Multiple contacts files found. Please keep only one contacts file in the script folder:")
        for c in existing:
            print(f" - {c.name}")
        raise FileNotFoundError("Multiple contacts files found.")
    else:
        extras = sorted(folder.glob("contacts*.xlsx"))
        if len(extras) == 1:
            contacts_file = extras[0]
            print(f"Using contacts file: {contacts_file.name}")
        elif len(extras) > 1:
            print("Multiple contacts files found. Please keep only one contacts file in the script folder:")
            for c in extras:
                print(f" - {c.name}")
            raise FileNotFoundError("Multiple contacts files found.")
        else:
            raise FileNotFoundError("contacts.xlsx not found in the script folder")

    wb = load_workbook(contacts_file)
    ws = wb.active

    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        name = str(row[0]).strip() if row[0] is not None else ""
        phone = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        message = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not name or not phone:
            continue

        contacts.append({
            "name": name,
            "phone": phone,
            "message": message
        })

    return contacts


def save_reports(results):
    today = datetime.now().strftime("%Y-%m-%d")
    json_file = BASE_DIR / f"whatsapp_report_{today}.json"
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=4, ensure_ascii=False)
    print(f"JSON Report Saved: {json_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = "WhatsApp Report"
    ws.append(["Name", "Phone", "Status", "Last 3 Messages"])
    for item in results:
        ws.append([
            item["Name"],
            item["Phone"],
            item["Status"],
            "\n".join(item["Last3Messages"])
        ])

    excel_file = BASE_DIR / f"whatsapp_report_{today}.xlsx"
    try:
        wb.save(excel_file)
        print(f"Excel Report Saved: {excel_file}")
    except PermissionError:
        alternate_file = BASE_DIR / f"whatsapp_report_{today}_{int(time.time())}.xlsx"
        wb.save(alternate_file)
        print(f"Excel Report save failed because the file was locked. Saved instead to: {alternate_file}")


def main():
    os.makedirs(PROFILE_DIR, exist_ok=True)
    os.makedirs("screenshots", exist_ok=True)

    try:
        contacts = read_contacts()
    except Exception as e:
        print(f"Error loading contacts: {e}")
        return

    if not contacts:
        print("No contacts found in contacts.xlsx.")
        return

    print(f"\nLoaded {len(contacts)} contact(s)\n")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("Playwright is not installed in this Python environment.")
        print(f'Install with: & "{sys.executable}" -m pip install playwright')
        print(f'Then run: & "{sys.executable}" -m playwright install chromium')
        return

    with sync_playwright() as p:
        PROFILE_DIR.mkdir(parents=True, exist_ok=True)
        SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)

        context = p.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=False,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"]
        )
        page = context.pages[0] if context.pages else context.new_page()

        print("Opening WhatsApp Web...")
        page.goto("https://web.whatsapp.com", wait_until="domcontentloaded", timeout=60000)

        try:
            page.wait_for_selector('[data-testid="chat-list"]', timeout=60000)
            print("WhatsApp Loaded Successfully")
        except:
            input("Scan QR Code and press ENTER...")

        results = []
        for contact in contacts:
            name = contact["name"]
            message = contact["message"].replace("{name}", name)
            status = "Failed"
            last_messages = []

            print(f"\nProcessing: {name}")

            try:
                search_box = find_search_box(page)
                if not search_box:
                    page.screenshot(path="search_error.png")
                    raise Exception("Search box not found")

                search_box.click()
                page.keyboard.press("Control+A")
                page.keyboard.press("Backspace")
                random_delay()

                try:
                    search_box.fill(name)
                except:
                    search_box.type(name)

                random_delay()

                # click the contact result explicitly
                contact_selector_candidates = [
                    f'span[title="{name}"]',
                    f'//span[text()="{name}"]',
                    f'//div[contains(@class, "_2wP_Y") and .//span[text()="{name}"]]',
                    f'//div[contains(@aria-label, "{name}")]',
                ]
                contact_clicked = False
                for sel in contact_selector_candidates:
                    try:
                        contact_element = page.locator(sel).first
                        contact_element.wait_for(state="visible", timeout=5000)
                        contact_element.click()
                        contact_clicked = True
                        break
                    except:
                        continue

                if not contact_clicked:
                    if contact.get("phone"):
                        print(f"Contact '{name}' not found; retrying search by phone {contact['phone']}")
                        search_box.click()
                        page.keyboard.press("Control+A")
                        page.keyboard.press("Backspace")
                        random_delay()
                        try:
                            search_box.fill(contact["phone"])
                        except:
                            search_box.type(contact["phone"])
                        random_delay()

                        for sel in contact_selector_candidates:
                            try:
                                contact_element = page.locator(sel).first
                                contact_element.wait_for(state="visible", timeout=5000)
                                contact_element.click()
                                contact_clicked = True
                                break
                            except:
                                continue

                if not contact_clicked:
                    print(f"Contact '{name}' not found in search results; pressing Enter as fallback")
                    page.keyboard.press("Enter")

                page.wait_for_timeout(3000)

                message_box = find_message_box(page)
                if not message_box:
                    page.screenshot(path=f"{SCREENSHOTS_DIR}/messagebox_error_{sanitize_filename(name)}.png")
                    raise Exception("Message box not found")

                message_box.click()
                random_delay()
                try:
                    message_box.fill(message)
                except:
                    message_box.type(message, delay=25)
                random_delay()
                page.keyboard.press("Enter")

                status = "Sent"
                print(f"Message Sent to {name}")

                screenshot_name = sanitize_filename(name) or "contact"
                screenshot_file = SCREENSHOTS_DIR / f"{screenshot_name}.png"
                page.screenshot(path=str(screenshot_file))
                print(f"Screenshot Saved: {screenshot_file}")

                try:
                    messages = page.locator("div.copyable-text")
                    count = messages.count()
                    for i in range(max(0, count - 3), count):
                        try:
                            text = messages.nth(i).inner_text()
                            if text.strip():
                                last_messages.append(text)
                        except:
                            pass
                except:
                    pass

            except Exception as e:
                print(f"Error with {name}: {e}")

            results.append({
                "Name": name,
                "Phone": contact["phone"],
                "Status": status,
                "Last3Messages": last_messages
            })
            random_delay()

        save_reports(results)
        print("\nAutomation Completed Successfully")
        context.close()


if __name__ == "__main__":
    main()
